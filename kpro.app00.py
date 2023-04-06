import streamlit as st
import os
import openpyxl
import datetime
import time

# 参照URL: https://www.youtube.com/watch?v=4nsTce1Oce8

#--------------------------------------------------------------------------
#Web画面の定義

st.image("k-pro_logo.png",width=400)
#st.title('K-Project')
st.header('「スキマプログラム」を活用して業務自動化・時短化を実現！')
st.subheader('「スキマプログラム」とは？')
st.markdown('スキマプログラムと個人的にネーミングしたもので、「隙間」を埋めるプログラムの意です。無料のプログラミングツールを活用して、今の作業環境に自動化ボタンを追加して時短化を図るプログラムです！K-Projectではプログラムの作成・保守と、ツール作成に関するアドバイス（人材育成）を行います。')
st.subheader('例えばこんなプログラム事例があります。')
st.markdown('・「伝票自動作成ツール」・・・Ｅｘｃｅｌで作成したリストを元に販売管理システム(ソフト)に自動転記して伝票を作成していきます。100枚でも1000枚でもＯＫ！主にインポート機能が使えないアナログシーンで絶大な効果を発揮します。')
#st.video("自動伝票作成プログラム.mp4")
st.markdown('・「賞状ファイルの自動作成ツール」・・・Ｅｘｃｅｌで作成した賞状のひな型に賞名、氏名、学校名、学年などの情報を自動入力し、ＰＤＦファイルに変換して一括印刷します。')
st.markdown('・「パート給与計算」・・・アプリを使用した打刻機能と給与明細を発行できるＥｘｃｅｌシートを連動させて毎月の給与計算を省力化。給与業務の担当者様がタイムカードや出勤簿を見てデータ入力している場合は入力作業自体が無くなります。')


#--------------------------------------------------------------------------
#エクセルファイルが無ければ作る
file_name = '新規のデータベース.xlsx'
if os.path.exists(file_name) == False:
    book = openpyxl.Workbook()
    book.save(file_name)

#エクセルファイルの一行目を設定
book = openpyxl.load_workbook(file_name)
sheet = book['Sheet']
sheet['A1'] = '氏名'
sheet['B1'] = '年齢'
sheet['C1'] = '住所'
sheet['D1'] = '連絡先①'
sheet['E1'] = 'メールアドレス①'
sheet['F1'] = '作成日'

book.save(file_name)


#--------------------------------------------------------------------------
#Webフォームの

with st.form(key = 'profile_form'):
    st.markdown('データ入力フォーム')
    name = st.text_input('名前')
    
    #age = st.text_input('年齢')
    age = st.slider(label='年齢',
                    min_value=0,
                    max_value=130,
                    value=30,
    )
    
    adress = st.text_input('住所')
    phone1 = st.text_input('連絡先①')
    mail = st.text_input('メールアドレス①')

    st.date_input('登録日')


    submit_btn = st.form_submit_button('送信')
    cancel_btn = st.form_submit_button('キャンセル')
    if submit_btn:
        st.text(f'入力された情報は {name},{age},{adress},{phone1},{mail} です')
        print({name},{age},{adress},{phone1},{mail})

#--------------------------------------------------------------------------
#エクセルファイルに書き込むために変数を整理
        

#時間を2020/1/1のようにスラッシュで区切る
        time = datetime.datetime.now()
        today = str(time.year) + '/' + str(time.month) + '/' + str(time.day) + '/' + str(time.hour)+ ':' + str(time.minute)+ '_' + str(time.second) 

#--------------------------------------------------------------------------
        book = openpyxl.load_workbook(file_name)
        row = sheet.max_row +1
        sheet = book['Sheet']
        
        sheet['A'+str(row)] = name
        sheet['B'+str(row)] = age
        sheet['C'+str(row)] = adress
        sheet['D'+str(row)] = phone1
        sheet['E'+str(row)] = mail
        sheet['F'+str(row)] = today
        book.save(file_name)
        book.close()

